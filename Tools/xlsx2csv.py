#! /usr/bin/env python
# -*- coding: utf-8 -*-

import xlrd
from openpyxl import load_workbook
import csv
import os,sys
import codecs

reload(sys)
sys.setdefaultencoding('UTF-8')


srcPath = "./../DesignData/"
dstPath = "./../Proj/Assets/Resources/Data/"

def xlsx_to_csv(fullname):
	name = os.path.splitext(fullname)[0]
	suffix = os.path.splitext(fullname)[1]
	if suffix != ".xlsx":
		return
	workbook = load_workbook(os.path.join(srcPath,fullname))
	for sheet in workbook.sheetnames:
		sheet_ranges = workbook[sheet]
		with codecs.open(os.path.join(dstPath,name+'.csv'), 'w', encoding='utf-8') as f:
			write = csv.writer(f)
			for row in sheet_ranges.rows:
				row_container = []
				for cell in row:
					if type(cell.value) == unicode:
						row_container.append(cell.value.encode('utf-8'))
					else:
						row_container.append(str(cell.value))

				write.writerow(row_container)
		break

	# table = workbook.sheet_by_index(0)
	# with codecs.open(os.path.join(dstPath,name+'.csv'), 'w', encoding='utf-8') as f:
	# 	write = csv.writer(f)
	# 	for row_num in range(table.nrows):
	# 		row_value = table.row_values(row_num)
	# 		write.writerow(row_value)
def convertAll():
	filelist = os.listdir(srcPath)
	for name in filelist:
		if not os.path.isdir(os.path.join(srcPath,name)):
			xlsx_to_csv(name)

if __name__ == '__main__':
	convertAll()




# -*- coding:utf-8 -*-

# from openpyxl import load_workbook
# import csv
# import os
# import sys

# def xlsx2csv(filename):
# 	try:
# 		#open the file you want to process
# 		xlsx_file_reader = load_workbook(filename = filename)

# 		#every sheet output to a csv file£¬filename is xlsx filename and sheet filename  '_' connection
# 		for sheet in xlsx_file_reader.get_sheet_names():
# 			#craete file with file name
# 			csv_filename = '{xlsx}_{sheet}.csv'.format(xlsx = os.path.splitext(filename.replace(' ', '_'))[0], sheet = sheet.replace(' ', '_'))

# 			#write csv file
# 			csv_file = file(csv_filename, 'wb')
# 			csv_file_writer = csv.writer(csv_file)

# 			#read the sheets of excel file
# 			sheet_ranges = xlsx_file_reader[sheet]

# 		   #Loop traversal every sheets' data
# 			for row in sheet_ranges.rows:
# 				row_container = []
# 				for cell in row:
# 					if type(cell.value) == unicode:
# 						row_container.append(cell.value.encode('utf-8'))
# 					else:
# 						row_container.append(str(cell.value))
# 				csv_file_writer.writerow(row_container)
# 			csv_file.close()
# 	except Exception as e:
# 		print (e)

# if __name__ == '__main__':
# 	xlsx2csv(r'/usr/testData/testfile1.xlsx')
# 	sys.exit(0)
